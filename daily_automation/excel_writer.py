# -*- coding: utf-8 -*-
# excel_writer.py

import csv, os, shutil, time
from datetime import datetime, date
from typing import List, Dict, Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# --- Config de tipado por nombre de columna ---
DATETIME_COL = "FECHA"
DATE_COL = "JORNADA"

NUMERIC_COLUMNS = {
    "CABIMPORTE","CABDESCUENTO","CABNETO","CANTIDAD","PRECIO","IVA",
    "IMPORTE","IMPORTESINIVA","DESCUENTO","IMPORTEDESCUENTO","IMPORTESINIVADESCUENTO",
    "TIENDA","COMENSALES","COSTE",
}
TEXT_FORCE_COLUMNS = {
    "IDTRANS","NSERIE","SERIE","NUMTIKET","NUMBARRA","NNUMBARRA",
    "NUMCLIE","PUNTOVENTA","NPUNTOVENTA","NUMCUEN","NNUMCUEN",
    "SERVICIO","NSERVICIO","ALMACEN","NALMACEN",
    "CAMARERO","NCAMARERO","MACROGRUPO","NMACROGRUPO","GRUPO","NGRUPO",
    "FAMILIA","NFAMILIA","TIPOPRODUCTO","NTIPOPRODUCTO",
    "PRODUCTO","NPRODUCTO","ANULADA","FORMATO","NFORMATO","ESTABLECIMIENTO",
    "CTACONTABLE","CECO","NIFCLIENTE","NOMBRECLIENTE","VENCIMIENTO","PROMOCION",
    "OBSERVACIONES","Turno","Denominacion 2","Factura","Motivo"
}

# ------------------------ Utilidades de parseo ------------------------

def safe_float(x: Any) -> Optional[float]:
    if x is None: 
        return None
    s = str(x).strip()
    if s == "" or s.lower() == "nan": 
        return None
    s = s.replace(" ", "").replace("\u00a0","").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

def parse_dt(s: str) -> Optional[datetime]:
    if not s: 
        return None
    s = s.strip().replace("Z","")
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S",
                "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%dT%H:%M", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s)
    except Exception:
        return None

def parse_d(s: str) -> Optional[date]:
    if not s: 
        return None
    s = s.strip().replace("Z","")
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None

def load_csv_rows(csv_path: str, encoding="utf-8") -> List[Dict[str, Any]]:
    print(f"📁 Cargando CSV: {csv_path}")
    start_time = time.time()
    
    with open(csv_path, newline="", encoding=encoding) as f:
        rows = list(csv.DictReader(f))
    
    elapsed = time.time() - start_time
    print(f"✅ CSV cargado: {len(rows)} filas en {elapsed:.2f}s")
    return rows

def read_header(ws) -> List[str]:
    return [
        (cell.value or "").strip() if isinstance(cell.value, str) else (cell.value or "")
        for cell in ws[1]
    ]

def coerce_value(col: str, val: Any):
    if val is None:
        return ""
    if isinstance(val, str) and val.strip().lower() == "none":
        val = ""
    if col in TEXT_FORCE_COLUMNS:
        return str(val)
    if col == DATETIME_COL:
        dt = parse_dt(str(val))
        return dt if dt else str(val)
    if col == DATE_COL:
        d = parse_d(str(val))
        return d if d else str(val)
    if col in NUMERIC_COLUMNS:
        num = safe_float(val)
        return num if num is not None else ("" if val == "" else str(val))
    return val

# ------------------------ Helpers de Excel ------------------------

def cell_has_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")

def cell_is_merged(cell) -> bool:
    return isinstance(cell, MergedCell)

def ensure_row_exists(ws, row_idx: int):
    """Asegura que la fila row_idx existe; si no, agrega filas hasta alcanzarla."""
    while ws.max_row < row_idx:
        ws.append([])

# ------------------------ Análisis previo para optimización ------------------------

def analyze_sheet_structure(ws, csv_cols, header):
    """
    Analiza la estructura de la hoja para optimizar el proceso de escritura
    """
    print(f"🔍 Analizando estructura de la hoja...")
    start_time = time.time()
    
    col_indices = {name: (header.index(name) + 1) for name in csv_cols}
    
    # Analizar celdas con fórmulas y merged en las columnas relevantes
    formula_cells = set()
    merged_cells = set()
    
    max_row_to_check = min(ws.max_row, 10000)  # Limitamos el análisis para no tardar demasiado
    
    for col_name, col_idx in col_indices.items():
        col_letter = get_column_letter(col_idx)
        formula_count = 0
        merged_count = 0
        
        for row_num in range(2, max_row_to_check + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            
            if cell_has_formula(cell):
                formula_cells.add((row_num, col_idx))
                formula_count += 1
            
            if cell_is_merged(cell):
                merged_cells.add((row_num, col_idx))
                merged_count += 1
        
        if formula_count > 0 or merged_count > 0:
            print(f"  📊 Columna {col_name}: {formula_count} fórmulas, {merged_count} merged")
    
    elapsed = time.time() - start_time
    print(f"✅ Análisis completado en {elapsed:.2f}s")
    print(f"  🔒 Total celdas con fórmula: {len(formula_cells)}")
    print(f"  🔗 Total celdas merged: {len(merged_cells)}")
    
    return formula_cells, merged_cells, col_indices

# ------------------------ Escritura optimizada ------------------------

def write_data_optimized(ws, rows, col_indices, formula_cells, merged_cells):
    """
    Escribe los datos de manera optimizada, evitando verificaciones innecesarias
    """
    print(f"✍️  Escribiendo {len(rows)} filas de datos...")
    start_time = time.time()
    
    first_data_row = 2
    last_progress = 0
    
    for i, record in enumerate(rows):
        r = first_data_row + i
        ensure_row_exists(ws, r)
        
        # Progreso cada 1000 filas
        if i > 0 and i % 1000 == 0:
            elapsed = time.time() - start_time
            progress = (i / len(rows)) * 100
            speed = i / elapsed if elapsed > 0 else 0
            eta = (len(rows) - i) / speed if speed > 0 else 0
            print(f"  📝 Progreso: {progress:.1f}% ({i:,}/{len(rows):,}) - {speed:.0f} filas/s - ETA: {eta:.1f}s")
        
        for col_name, cidx in col_indices.items():
            # Verificación optimizada: solo verificar si está en los sets precalculados
            if (r, cidx) in formula_cells or (r, cidx) in merged_cells:
                continue
            
            cell = ws.cell(row=r, column=cidx)
            cell.value = coerce_value(col_name, record.get(col_name, ""))
    
    elapsed = time.time() - start_time
    speed = len(rows) / elapsed if elapsed > 0 else 0
    print(f"✅ Datos escritos en {elapsed:.2f}s ({speed:.0f} filas/s)")

def clean_old_data_optimized(ws, last_new_row, col_indices, formula_cells, merged_cells):
    """
    Limpia datos antiguos de manera optimizada
    """
    last_row = ws.max_row
    if last_row <= last_new_row:
        return
    
    rows_to_clean = last_row - last_new_row
    print(f"🧹 Limpiando {rows_to_clean} filas antiguas...")
    start_time = time.time()
    
    cleaned_count = 0
    for r in range(last_new_row + 1, last_row + 1):
        for col_name, cidx in col_indices.items():
            if (r, cidx) not in formula_cells and (r, cidx) not in merged_cells:
                cell = ws.cell(row=r, column=cidx)
                cell.value = ""
                cleaned_count += 1
    
    elapsed = time.time() - start_time
    print(f"✅ Limpieza completada en {elapsed:.2f}s ({cleaned_count} celdas limpiadas)")

def apply_date_formatting(ws, header, first_data_row, last_new_row):
    """
    Aplica formato de fecha de manera optimizada
    """
    print(f"📅 Aplicando formatos de fecha...")
    start_time = time.time()
    
    formatted_count = 0
    
    # Formato DATETIME_COL
    try:
        idx = header.index(DATETIME_COL) + 1
        col_letter = get_column_letter(idx)
        for row_num in range(first_data_row, last_new_row + 1):
            cell = ws.cell(row=row_num, column=idx)
            if not cell_is_merged(cell) and isinstance(cell.value, datetime):
                cell.number_format = "dd/mm/yyyy hh:mm"
                formatted_count += 1
    except ValueError:
        pass
    
    # Formato DATE_COL
    try:
        idx = header.index(DATE_COL) + 1
        col_letter = get_column_letter(idx)
        for row_num in range(first_data_row, last_new_row + 1):
            cell = ws.cell(row=row_num, column=idx)
            if (not cell_is_merged(cell) and 
                isinstance(cell.value, date) and 
                not isinstance(cell.value, datetime)):
                cell.number_format = "dd/mm/yyyy"
                formatted_count += 1
    except ValueError:
        pass
    
    elapsed = time.time() - start_time
    print(f"✅ Formatos aplicados en {elapsed:.2f}s ({formatted_count} celdas)")

# ------------------------ Función principal optimizada ------------------------

def overwrite_non_formula_cells_with_csv(xlsx_path: str, sheet_name: str, csv_path: str, backup=True):
    """
    Versión optimizada con logs detallados y análisis previo
    """
    print(f"\n🚀 INICIANDO PROCESO DE ESCRITURA EXCEL")
    print(f"📄 Archivo: {xlsx_path}")
    print(f"📋 Hoja: {sheet_name}")
    print(f"📊 CSV: {csv_path}")
    print("="*50)
    
    total_start_time = time.time()
    
    # Backup
    if backup:
        print(f"💾 Creando backup...")
        backup_start = time.time()
        base, ext = os.path.splitext(xlsx_path)
        bk = f"{base}.backup{ext}"
        shutil.copy2(xlsx_path, bk)
        backup_time = time.time() - backup_start
        print(f"✅ Backup creado en {backup_time:.2f}s: {bk}")

    # 0) Cargar datos CSV
    rows = load_csv_rows(csv_path)

    # 1) Abrir libro y hoja
    print(f"📖 Abriendo archivo Excel...")
    excel_start = time.time()
    wb = load_workbook(xlsx_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"No existe la hoja '{sheet_name}'")
    ws = wb[sheet_name]
    excel_time = time.time() - excel_start
    print(f"✅ Excel abierto en {excel_time:.2f}s")

    # 2) Cabecera de hoja
    header = [str(h) for h in read_header(ws)]
    if not header or all(h == "" for h in header):
        raise SystemExit("Cabecera de la hoja vacía")
    print(f"📋 Cabeceras encontradas: {len(header)}")

    # 3) Columnas a escribir
    if rows:
        csv_cols = [c for c in rows[0].keys() if c in header]
    else:
        csv_cols = []
    
    if not csv_cols:
        wb.save(xlsx_path)
        print("❌ No hay columnas válidas en el CSV. No se realizaron cambios.")
        return
    
    print(f"📊 Columnas a procesar: {len(csv_cols)}")
    print(f"📏 Filas máximas en hoja: {ws.max_row}")

    # 4) Análisis optimizado de estructura
    formula_cells, merged_cells, col_indices = analyze_sheet_structure(ws, csv_cols, header)

    # 5) Escritura optimizada
    write_data_optimized(ws, rows, col_indices, formula_cells, merged_cells)

    # 6) Limpieza optimizada
    last_new_row = 2 + len(rows) - 1 if rows else 1
    clean_old_data_optimized(ws, last_new_row, col_indices, formula_cells, merged_cells)

    # 7) Formatos de fecha
    if rows:
        apply_date_formatting(ws, header, 2, last_new_row)

    # 8) Guardar
    print(f"💾 Guardando archivo...")
    save_start = time.time()
    wb.save(xlsx_path)
    save_time = time.time() - save_start
    print(f"✅ Archivo guardado en {save_time:.2f}s")

    # Resumen final
    total_time = time.time() - total_start_time
    print("="*50)
    print(f"🎉 PROCESO COMPLETADO")
    print(f"⏱️  Tiempo total: {total_time:.2f}s")
    print(f"📊 Filas procesadas: {len(rows):,}")
    print(f"📈 Velocidad promedio: {len(rows)/total_time:.0f} filas/s")
    print(f"🔒 Celdas con fórmula protegidas: {len(formula_cells):,}")
    print(f"🔗 Celdas merged protegidas: {len(merged_cells):,}")
    print("="*50)